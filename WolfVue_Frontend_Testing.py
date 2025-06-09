#!/usr/bin/env python3
"""
WolfVue Desktop GUI: Wildlife Video Classifier
A modern, customizable desktop interface for the WolfVue wildlife classification system.

Created by Nathan Bluto
Data from The Gray Wolf Research Project
Facilitated by Dr. Ausband

==============================================================================
CUSTOMIZATION GUIDE:
==============================================================================

1. VISUAL CUSTOMIZATION:
   - Modify the THEME_CONFIG dictionary below to change colors, fonts, and spacing
   - Adjust LAYOUT_CONFIG for window sizes and proportions
   - Change ANIMATION_CONFIG for different visual effects

2. FUNCTIONALITY CUSTOMIZATION:
   - Modify DEFAULT_PROCESSING_SETTINGS for default algorithm parameters
   - Adjust EXPORT_FORMATS to add/remove export options
   - Change SUPPORTED_FILE_TYPES to support additional file formats

3. UI BEHAVIOR CUSTOMIZATION:
   - Modify UPDATE_INTERVALS for different refresh rates
   - Adjust VALIDATION_RULES for input validation
   - Change DIALOG_SETTINGS for popup behavior

4. ADVANCED CUSTOMIZATION:
   - Override create_custom_widgets() to add new interface elements
   - Modify the StyleManager class to create new themes
   - Extend the ExportManager class for new export formats

==============================================================================
"""

import sys
import os
import json
import yaml
import time
import shutil
import csv
import random
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, asdict
from enum import Enum

# PyQt6 imports
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
    QGridLayout, QLabel, QPushButton, QLineEdit, QTextEdit, QProgressBar,
    QFileDialog, QTabWidget, QGroupBox, QSpinBox, QDoubleSpinBox,
    QCheckBox, QComboBox, QSplitter, QFrame, QScrollArea, QTableWidget,
    QTableWidgetItem, QHeaderView, QMessageBox, QDialog, QDialogButtonBox,
    QFormLayout, QSlider, QStatusBar, QMenuBar, QMenu, QToolBar, QListWidget,
    QListWidgetItem, QAbstractItemView, QColorDialog, QFontDialog
)
from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, QTimer, QSettings, QSize, QRect, pyqtSlot,
    QMimeData, QThreadPool, QRunnable, QObject
)
from PyQt6.QtGui import (
    QFont, QIcon, QPalette, QColor, QPixmap, QPainter, QBrush, QLinearGradient,
    QAction, QFontMetrics, QClipboard, QKeySequence, QShortcut
)

# ============================================================================
# CUSTOMIZATION CONFIGURATION - MODIFY THESE TO CUSTOMIZE THE INTERFACE
# ============================================================================

# Theme and visual configuration - Easy to modify for different looks
THEME_CONFIG = {
    'primary_color': '#4CAF50',        # Main theme color (green)
    'primary_dark': '#45a049',         # Darker shade of primary
    'primary_light': '#5CBF60',        # Lighter shade of primary
    'secondary_color': '#2196F3',      # Secondary color (blue)
    'background_color': '#f8f9fa',     # Main background
    'surface_color': '#ffffff',        # Card/panel backgrounds
    'border_color': '#dee2e6',         # Border colors
    'text_primary': '#495057',         # Main text color
    'text_secondary': '#6c757d',       # Secondary text color
    'text_muted': '#adb5bd',          # Muted text color
    'success_color': '#28a745',        # Success messages
    'warning_color': '#ffc107',        # Warning messages
    'error_color': '#dc3545',          # Error messages
    'info_color': '#17a2b8',          # Info messages
}

# Font configuration - Adjust for different font preferences
FONT_CONFIG = {
    'main_font_family': 'Segoe UI',    # Main font (Windows default)
    'main_font_size': 13,              # Base font size
    'title_font_size': 28,             # Large title size
    'subtitle_font_size': 14,          # Subtitle size
    'small_font_size': 11,             # Small text size
    'monospace_font': 'Consolas',      # Monospace font for technical data
}

# Layout and spacing configuration
LAYOUT_CONFIG = {
    'window_min_width': 1200,          # Minimum window width
    'window_min_height': 800,          # Minimum window height
    'left_panel_max_width': 450,       # Configuration panel width
    'splitter_ratio': [400, 800],      # Left:Right panel ratio
    'border_radius': 8,                # Corner radius for elements
    'spacing_small': 8,                # Small spacing
    'spacing_medium': 16,              # Medium spacing
    'spacing_large': 24,               # Large spacing
    'padding_small': 8,                # Small padding
    'padding_medium': 12,              # Medium padding
    'padding_large': 20,               # Large padding
}

# Animation and interaction configuration
ANIMATION_CONFIG = {
    'enable_animations': True,         # Enable/disable animations
    'hover_transition_time': 200,      # Button hover animation time (ms)
    'progress_update_interval': 100,   # Progress bar update frequency (ms)
    'status_display_time': 3000,       # Status message display time (ms)
}

# Default processing settings - Modify these for different default behaviors
DEFAULT_PROCESSING_SETTINGS = {
    'confidence_threshold': 0.40,
    'dominant_species_threshold': 0.9,
    'max_species_transitions': 5,
    'consecutive_empty_frames': 15,
    'image_confidence_threshold': 0.65,
    'image_min_detections': 1,
    'image_multi_species_threshold': 0.60,
    'image_unsorted_min_confidence': 0.35,
    'image_unsorted_max_confidence': 0.65,
}

# Supported file types - Add/remove file extensions as needed
SUPPORTED_FILE_TYPES = {
    'video_extensions': ['.mp4', '.avi', '.mov', '.mkv', '.wmv', '.flv', '.webm'],
    'image_extensions': ['.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif', '.webp'],
    'model_extensions': ['.pt', '.pth', '.onnx'],
    'config_extensions': ['.yaml', '.yml', '.json'],
}

# Export formats configuration - Add new export options here
EXPORT_FORMATS = {
    'csv': {'name': 'CSV (Comma Separated)', 'extension': '.csv'},
    'excel': {'name': 'Excel Spreadsheet', 'extension': '.xlsx'},
    'json': {'name': 'JSON Data', 'extension': '.json'},
    'txt': {'name': 'Text Report', 'extension': '.txt'},
}

# Update intervals configuration
UPDATE_INTERVALS = {
    'progress_update_ms': 100,         # How often to update progress
    'status_refresh_ms': 500,          # Status message refresh rate
    'auto_save_minutes': 5,            # Auto-save settings interval
}

# Validation rules configuration
VALIDATION_RULES = {
    'min_confidence': 0.01,            # Minimum confidence threshold
    'max_confidence': 0.99,            # Maximum confidence threshold
    'min_detections': 1,               # Minimum detections required
    'max_detections': 100,             # Maximum detections allowed
    'max_path_length': 260,            # Maximum file path length
}

# Dialog and popup settings
DIALOG_SETTINGS = {
    'auto_close_success': True,        # Auto-close success dialogs
    'auto_close_delay': 2000,          # Auto-close delay (ms)
    'confirm_destructive_actions': True, # Confirm before destructive actions
    'remember_dialog_positions': True,  # Remember dialog positions
}

# ============================================================================
# END CUSTOMIZATION CONFIGURATION
# ============================================================================

# Try to import optional dependencies for enhanced functionality
try:
    from ultralytics import YOLO
    import cv2
    YOLO_AVAILABLE = True
except ImportError:
    YOLO_AVAILABLE = False
    print("Warning: YOLO dependencies not available. Running in demo mode.")

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: Excel export not available. Install openpyxl for Excel support.")

@dataclass
class ProcessingSettings:
    """
    Configuration settings for processing.
    
    This class holds all the settings that control how videos and images
    are processed. You can modify the DEFAULT_PROCESSING_SETTINGS dict
    above to change the default values.
    """
    input_folder: str = ""
    output_folder: str = ""
    model_path: str = ""
    config_file: str = ""
    
    # Video processing parameters
    confidence_threshold: float = DEFAULT_PROCESSING_SETTINGS['confidence_threshold']
    dominant_species_threshold: float = DEFAULT_PROCESSING_SETTINGS['dominant_species_threshold']
    max_species_transitions: int = DEFAULT_PROCESSING_SETTINGS['max_species_transitions']
    consecutive_empty_frames: int = DEFAULT_PROCESSING_SETTINGS['consecutive_empty_frames']
    
    # Image processing parameters
    image_confidence_threshold: float = DEFAULT_PROCESSING_SETTINGS['image_confidence_threshold']
    image_min_detections: int = DEFAULT_PROCESSING_SETTINGS['image_min_detections']
    image_multi_species_threshold: float = DEFAULT_PROCESSING_SETTINGS['image_multi_species_threshold']
    image_unsorted_min_confidence: float = DEFAULT_PROCESSING_SETTINGS['image_unsorted_min_confidence']
    image_unsorted_max_confidence: float = DEFAULT_PROCESSING_SETTINGS['image_unsorted_max_confidence']

class ThemeManager:
    """
    Manages application themes and styling.
    
    To create a new theme:
    1. Copy THEME_CONFIG and modify colors
    2. Call apply_theme() with your new config
    3. Save as preset using save_theme_preset()
    """
    
    def __init__(self):
        self.current_theme = THEME_CONFIG.copy()
        self.font_config = FONT_CONFIG.copy()
        
    def apply_theme(self, theme_config: dict = None):
        """Apply a theme configuration to generate CSS styles."""
        if theme_config:
            self.current_theme.update(theme_config)
            
        return self.generate_stylesheet()
    
    def generate_stylesheet(self) -> str:
        """Generate the complete CSS stylesheet from theme configuration."""
        theme = self.current_theme
        font = self.font_config
        layout = LAYOUT_CONFIG
        
        return f"""
        /* ============================================
           MAIN APPLICATION STYLING
           ============================================ */
        QMainWindow {{
            background-color: {theme['background_color']};
            font-family: '{font['main_font_family']}';
            font-size: {font['main_font_size']}px;
        }}
        
        /* ============================================
           GROUP BOXES AND CONTAINERS
           ============================================ */
        QGroupBox {{
            font-weight: bold;
            border: 2px solid {theme['border_color']};
            border-radius: {layout['border_radius']}px;
            margin: {layout['spacing_small']}px 0px;
            padding-top: {layout['spacing_small']}px;
            background-color: {theme['surface_color']};
            color: {theme['text_primary']};
        }}
        
        QGroupBox::title {{
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 {layout['padding_small']}px 0 {layout['padding_small']}px;
            color: {theme['text_primary']};
        }}
        
        /* ============================================
           INPUT FIELDS
           ============================================ */
        QLineEdit {{
            border: 2px solid {theme['border_color']};
            border-radius: 6px;
            padding: {layout['padding_small']}px {layout['padding_medium']}px;
            font-size: {font['main_font_size']}px;
            background-color: {theme['surface_color']};
            color: {theme['text_primary']};
        }}
        
        QLineEdit:focus {{
            border-color: {theme['primary_color']};
        }}
        
        QLineEdit:disabled {{
            background-color: {theme['background_color']};
            color: {theme['text_muted']};
        }}
        
        /* ============================================
           LABELS AND TEXT
           ============================================ */
        QLabel {{
            color: {theme['text_primary']};
            font-size: {font['main_font_size']}px;
        }}
        
        /* ============================================
           TABLES
           ============================================ */
        QTableWidget {{
            background-color: {theme['surface_color']};
            border: 1px solid {theme['border_color']};
            border-radius: {layout['border_radius']}px;
            gridline-color: {theme['border_color']};
            color: {theme['text_primary']};
        }}
        
        QTableWidget::item {{
            padding: {layout['padding_small']}px;
            border-bottom: 1px solid {theme['border_color']};
        }}
        
        QTableWidget::item:selected {{
            background-color: {theme['primary_color']};
            color: white;
        }}
        
        QHeaderView::section {{
            background-color: {theme['background_color']};
            padding: {layout['padding_medium']}px;
            border: 1px solid {theme['border_color']};
            font-weight: bold;
            color: {theme['text_primary']};
        }}
        
        /* ============================================
           PROGRESS BARS
           ============================================ */
        QProgressBar {{
            border: 2px solid {theme['border_color']};
            border-radius: {layout['border_radius']}px;
            text-align: center;
            background-color: {theme['background_color']};
            color: {theme['text_primary']};
            font-weight: bold;
        }}
        
        QProgressBar::chunk {{
            background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                stop:0 {theme['primary_color']}, stop:1 {theme['primary_light']});
            border-radius: {layout['border_radius']-2}px;
        }}
        
        /* ============================================
           SLIDERS
           ============================================ */
        QSlider::groove:horizontal {{
            border: 1px solid {theme['border_color']};
            height: 8px;
            background: {theme['background_color']};
            border-radius: 4px;
        }}
        
        QSlider::handle:horizontal {{
            background: {theme['primary_color']};
            border: 2px solid {theme['primary_dark']};
            width: 18px;
            margin: -5px 0;
            border-radius: 9px;
        }}
        
        QSlider::sub-page:horizontal {{
            background: {theme['primary_color']};
            border-radius: 4px;
        }}
        
        /* ============================================
           COMBO BOXES
           ============================================ */
        QComboBox {{
            border: 2px solid {theme['border_color']};
            border-radius: 6px;
            padding: {layout['padding_small']}px {layout['padding_medium']}px;
            background-color: {theme['surface_color']};
            color: {theme['text_primary']};
        }}
        
        QComboBox:focus {{
            border-color: {theme['primary_color']};
        }}
        
        QComboBox::drop-down {{
            border: none;
            width: 20px;
        }}
        
        QComboBox::down-arrow {{
            width: 12px;
            height: 12px;
        }}
        
        /* ============================================
           SCROLL BARS
           ============================================ */
        QScrollBar:vertical {{
            border: none;
            background: {theme['background_color']};
            width: 14px;
            border-radius: 7px;
        }}
        
        QScrollBar::handle:vertical {{
            background: {theme['border_color']};
            border-radius: 7px;
            min-height: 20px;
        }}
        
        QScrollBar::handle:vertical:hover {{
            background: {theme['text_muted']};
        }}
        
        /* ============================================
           MENU AND STATUS BAR
           ============================================ */
        QMenuBar {{
            background-color: {theme['surface_color']};
            color: {theme['text_primary']};
            border-bottom: 1px solid {theme['border_color']};
        }}
        
        QMenuBar::item {{
            padding: {layout['padding_small']}px {layout['padding_medium']}px;
        }}
        
        QMenuBar::item:selected {{
            background-color: {theme['primary_color']};
            color: white;
        }}
        
        QStatusBar {{
            background-color: {theme['surface_color']};
            color: {theme['text_secondary']};
            border-top: 1px solid {theme['border_color']};
        }}
        """
    
    def save_theme_preset(self, name: str):
        """Save current theme as a preset."""
        settings = QSettings()
        settings.setValue(f"themes/{name}", json.dumps(self.current_theme))
    
    def load_theme_preset(self, name: str) -> bool:
        """Load a saved theme preset."""
        settings = QSettings()
        theme_data = settings.value(f"themes/{name}")
        if theme_data:
            try:
                self.current_theme = json.loads(theme_data)
                return True
            except json.JSONDecodeError:
                return False
        return False

class ModernButton(QPushButton):
    """
    Custom styled button with modern appearance and configurable styling.
    
    Usage:
        primary_btn = ModernButton("Start", primary=True)
        secondary_btn = ModernButton("Cancel", button_type="secondary")
        danger_btn = ModernButton("Delete", button_type="danger")
    """
    
    def __init__(self, text="", icon=None, primary=False, button_type="default", parent=None):
        super().__init__(text, parent)
        self.primary = primary
        self.button_type = button_type
        self.setupStyle()
        
        # Add keyboard shortcut support
        if icon:
            self.setIcon(icon)
            
    def setupStyle(self):
        """Setup the button styling based on type and theme."""
        theme = THEME_CONFIG
        layout = LAYOUT_CONFIG
        
        if self.primary or self.button_type == "primary":
            # Primary button styling (main actions)
            self.setStyleSheet(f"""
                QPushButton {{
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 {theme['primary_color']}, stop:1 {theme['primary_dark']});
                    border: none;
                    border-radius: {layout['border_radius']}px;
                    color: white;
                    font-weight: bold;
                    padding: {layout['padding_medium']}px {layout['padding_large']}px;
                    font-size: 14px;
                    min-height: 20px;
                }}
                QPushButton:hover {{
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 {theme['primary_light']}, stop:1 {theme['primary_color']});
                }}
                QPushButton:pressed {{
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 {theme['primary_dark']}, stop:1 #357a38);
                }}
                QPushButton:disabled {{
                    background: {theme['text_muted']};
                    color: #666666;
                }}
            """)
        elif self.button_type == "danger":
            # Danger button styling (destructive actions)
            self.setStyleSheet(f"""
                QPushButton {{
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 {theme['error_color']}, stop:1 #c82333);
                    border: none;
                    border-radius: {layout['border_radius']}px;
                    color: white;
                    font-weight: bold;
                    padding: {layout['padding_small']}px {layout['padding_medium']}px;
                    font-size: 13px;
                }}
                QPushButton:hover {{
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #e04858, stop:1 {theme['error_color']});
                }}
                QPushButton:pressed {{
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #c82333, stop:1 #a71e2a);
                }}
            """)
        elif self.button_type == "success":
            # Success button styling (positive actions)
            self.setStyleSheet(f"""
                QPushButton {{
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 {theme['success_color']}, stop:1 #218838);
                    border: none;
                    border-radius: {layout['border_radius']}px;
                    color: white;
                    font-weight: 500;
                    padding: {layout['padding_small']}px {layout['padding_medium']}px;
                    font-size: 13px;
                }}
                QPushButton:hover {{
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #34ce57, stop:1 {theme['success_color']});
                }}
            """)
        else:
            # Default/secondary button styling
            self.setStyleSheet(f"""
                QPushButton {{
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 {theme['surface_color']}, stop:1 #e9ecef);
                    border: 2px solid {theme['border_color']};
                    border-radius: {layout['border_radius']}px;
                    color: {theme['text_primary']};
                    font-weight: 500;
                    padding: {layout['padding_small']}px {layout['padding_medium']}px;
                    font-size: 13px;
                }}
                QPushButton:hover {{
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 #e9ecef, stop:1 {theme['border_color']});
                    border-color: {theme['text_muted']};
                }}
                QPushButton:pressed {{
                    background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                        stop:0 {theme['border_color']}, stop:1 #ced4da);
                }}
                QPushButton:disabled {{
                    background: {theme['surface_color']};
                    border-color: #e9ecef;
                    color: {theme['text_muted']};
                }}
            """)

class ExportManager:
    """
    Handles exporting data to various formats.
    
    To add a new export format:
    1. Add the format to EXPORT_FORMATS configuration
    2. Add a method here following the pattern export_to_[format]()
    3. Register it in the export_data() method
    """
    
    def __init__(self):
        self.supported_formats = EXPORT_FORMATS
        
    def export_data(self, data: List[dict], file_path: str, format_type: str) -> bool:
        """
        Export data to the specified format.
        
        Args:
            data: List of result dictionaries to export
            file_path: Output file path
            format_type: Export format ('csv', 'excel', 'json', 'txt')
            
        Returns:
            bool: True if export successful, False otherwise
        """
        try:
            if format_type == 'csv':
                return self.export_to_csv(data, file_path)
            elif format_type == 'excel':
                return self.export_to_excel(data, file_path)
            elif format_type == 'json':
                return self.export_to_json(data, file_path)
            elif format_type == 'txt':
                return self.export_to_txt(data, file_path)
            else:
                raise ValueError(f"Unsupported export format: {format_type}")
                
        except Exception as e:
            print(f"Export error: {e}")
            return False
    
    def export_to_csv(self, data: List[dict], file_path: str) -> bool:
        """Export data to CSV format."""
        if not data:
            return False
            
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['filename', 'classification', 'confidence', 'file_type', 
                         'processing_time', 'timestamp']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for row in data:
                # Add timestamp if not present
                if 'timestamp' not in row:
                    row['timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                writer.writerow(row)
        return True
    
    def export_to_excel(self, data: List[dict], file_path: str) -> bool:
        """Export data to Excel format with formatting."""
        if not EXCEL_AVAILABLE:
            raise ImportError("Excel export requires openpyxl. Install with: pip install openpyxl")
            
        if not data:
            return False
            
        # Convert to DataFrame for easier Excel export
        df = pd.DataFrame(data)
        
        # Add timestamp column if not present
        if 'timestamp' not in df.columns:
            df['timestamp'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # Create Excel file with formatting
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Results', index=False)
            
            # Create summary sheet
            summary_data = self.create_summary_data(data)
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format the worksheets
            workbook = writer.book
            self.format_excel_sheets(workbook)
            
        return True
    
    def export_to_json(self, data: List[dict], file_path: str) -> bool:
        """Export data to JSON format."""
        export_data = {
            'metadata': {
                'export_time': datetime.now().isoformat(),
                'total_files': len(data),
                'export_format': 'json'
            },
            'results': data,
            'summary': self.create_summary_data(data)
        }
        
        with open(file_path, 'w', encoding='utf-8') as jsonfile:
            json.dump(export_data, jsonfile, indent=2, ensure_ascii=False)
        return True
    
    def export_to_txt(self, data: List[dict], file_path: str) -> bool:
        """Export data to formatted text report."""
        with open(file_path, 'w', encoding='utf-8') as txtfile:
            txtfile.write("WolfVue Processing Report\n")
            txtfile.write("=" * 50 + "\n\n")
            txtfile.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            txtfile.write(f"Total Files Processed: {len(data)}\n\n")
            
            # Summary statistics
            summary = self.create_summary_data(data)
            txtfile.write("Summary Statistics:\n")
            txtfile.write("-" * 20 + "\n")
            for item in summary:
                txtfile.write(f"{item['metric']}: {item['value']}\n")
            
            txtfile.write("\n\nDetailed Results:\n")
            txtfile.write("-" * 20 + "\n")
            
            # Detailed results
            for i, result in enumerate(data, 1):
                txtfile.write(f"\n{i}. {result.get('filename', 'Unknown')}\n")
                txtfile.write(f"   Classification: {result.get('classification', 'Unknown')}\n")
                txtfile.write(f"   Confidence: {result.get('confidence', 0):.3f}\n")
                txtfile.write(f"   File Type: {result.get('file_type', 'Unknown')}\n")
                txtfile.write(f"   Processing Time: {result.get('processing_time', 0):.2f}s\n")
                
        return True
    
    def create_summary_data(self, data: List[dict]) -> List[dict]:
        """Create summary statistics from the data."""
        if not data:
            return []
            
        # Count by classification
        classifications = {}
        file_types = {}
        total_confidence = 0
        total_processing_time = 0
        
        for result in data:
            # Count classifications
            classification = result.get('classification', 'Unknown')
            classifications[classification] = classifications.get(classification, 0) + 1
            
            # Count file types
            file_type = result.get('file_type', 'Unknown')
            file_types[file_type] = file_types.get(file_type, 0) + 1
            
            # Sum for averages
            total_confidence += result.get('confidence', 0)
            total_processing_time += result.get('processing_time', 0)
        
        summary = [
            {'metric': 'Total Files', 'value': len(data)},
            {'metric': 'Average Confidence', 'value': f"{total_confidence/len(data):.3f}"},
            {'metric': 'Total Processing Time', 'value': f"{total_processing_time:.2f}s"},
            {'metric': 'Average Processing Time', 'value': f"{total_processing_time/len(data):.2f}s"},
        ]
        
        # Add classification counts
        for classification, count in sorted(classifications.items()):
            percentage = (count / len(data)) * 100
            summary.append({
                'metric': f'Classification: {classification}', 
                'value': f"{count} ({percentage:.1f}%)"
            })
            
        return summary
    
    def format_excel_sheets(self, workbook):
        """Apply formatting to Excel sheets."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Format Results sheet
            if 'Results' in workbook.sheetnames:
                ws = workbook['Results']
                
                # Header formatting
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                    
        except ImportError:
            pass  # Skip formatting if openpyxl styles not available
    
    def copy_to_clipboard(self, data: List[dict], format_type: str = 'tab_separated') -> bool:
        """
        Copy data to clipboard in various formats.
        
        Args:
            data: Data to copy
            format_type: 'tab_separated', 'csv', 'json'
        """
        try:
            clipboard = QApplication.clipboard()
            
            if format_type == 'tab_separated':
                # Format for Excel/spreadsheet pasting
                if not data:
                    return False
                    
                # Create header row
                headers = ['Filename', 'Classification', 'Confidence', 'File Type', 'Processing Time']
                content = '\t'.join(headers) + '\n'
                
                # Add data rows
                for row in data:
                    values = [
                        str(row.get('filename', '')),
                        str(row.get('classification', '')),
                        f"{row.get('confidence', 0):.3f}",
                        str(row.get('file_type', '')),
                        f"{row.get('processing_time', 0):.2f}"
                    ]
                    content += '\t'.join(values) + '\n'
                    
                clipboard.setText(content)
                
            elif format_type == 'csv':
                # CSV format
                content = "Filename,Classification,Confidence,File Type,Processing Time\n"
                for row in data:
                    values = [
                        f'"{row.get("filename", "")}"',
                        f'"{row.get("classification", "")}"',
                        f"{row.get('confidence', 0):.3f}",
                        f'"{row.get("file_type", "")}"',
                        f"{row.get('processing_time', 0):.2f}"
                    ]
                    content += ','.join(values) + '\n'
                clipboard.setText(content)
                
            elif format_type == 'json':
                # JSON format
                clipboard.setText(json.dumps(data, indent=2))
                
            return True
            
        except Exception as e:
            print(f"Clipboard copy error: {e}")
            return False

class ProcessingWorker(QThread):
    """
    Worker thread for processing videos and images.
    
    This runs in a separate thread to prevent the UI from freezing during
    long processing operations. It emits signals to update the main thread.
    
    Signals:
        progress_updated: Emitted with (progress_percent, status_message)
        file_processed: Emitted when a single file is completed
        processing_complete: Emitted when all processing is done
        error_occurred: Emitted when an error happens
    """
    
    # Define signals for communication with main thread
    progress_updated = pyqtSignal(int, str)  # progress percentage, status message
    file_processed = pyqtSignal(dict)        # result data for one file
    processing_complete = pyqtSignal(list)   # all results when complete
    error_occurred = pyqtSignal(str)         # error message
    
    def __init__(self, settings: ProcessingSettings):
        super().__init__()
        self.settings = settings
        self.is_cancelled = False
        
    def run(self):
        """
        Main processing loop that runs in the background thread.
        
        This method should contain all the actual YOLO processing logic
        from your original script. For now, it simulates the processing.
        """
        try:
            if YOLO_AVAILABLE:
                # Use real YOLO processing
                self.process_with_yolo()
            else:
                # Use simulation for demo
                self.simulate_processing()
                
        except Exception as e:
            self.error_occurred.emit(str(e))
    
    def process_with_yolo(self):
        """Process files using actual YOLO model (when available)."""
        # This would contain the actual processing logic from your original script
        # For now, fall back to simulation
        self.simulate_processing()
    
    def simulate_processing(self):
        """
        Simulate the processing for demo purposes.
        
        This creates fake results that look like real processing output.
        Replace this with actual YOLO processing logic.
        """
        # Simulate finding files
        self.progress_updated.emit(5, "Scanning for video and image files...")
        time.sleep(0.5)
        
        # Simulate file discovery based on actual folder contents if possible
        demo_files = []
        
        if os.path.exists(self.settings.input_folder):
            # Get real files from the input folder
            input_path = Path(self.settings.input_folder)
            video_exts = SUPPORTED_FILE_TYPES['video_extensions']
            image_exts = SUPPORTED_FILE_TYPES['image_extensions']
            
            for ext in video_exts + image_exts:
                demo_files.extend(list(input_path.glob(f'*{ext}')))
                demo_files.extend(list(input_path.glob(f'*{ext.upper()}')))
        
        # If no real files found, use demo filenames
        if not demo_files:
            demo_files = [
                "camera_001_video.mp4", "trail_cam_002.jpg", "wildlife_003.avi",
                "nature_004.png", "forest_005.mov", "deer_006.jpg", "wolf_007.mp4",
                "bear_008.jpg", "elk_009.avi", "fox_010.png"
            ]
        
        total_files = len(demo_files)
        results = []
        
        # Simulate processing each file
        for i, filename in enumerate(demo_files):
            if self.is_cancelled:
                return
                
            # Calculate progress (5% for scanning + 90% for processing + 5% for cleanup)
            progress = int((i + 1) / total_files * 90) + 5
            self.progress_updated.emit(progress, f"Processing {Path(filename).name}...")
            
            # Simulate processing time based on file type
            if isinstance(filename, Path):
                filename = filename.name
                
            if any(filename.lower().endswith(ext) for ext in SUPPORTED_FILE_TYPES['video_extensions']):
                time.sleep(0.8)  # Videos take longer
                file_type = 'video'
                processing_time = random.uniform(2.5, 8.2)
            else:
                time.sleep(0.3)  # Images are faster
                file_type = 'image'
                processing_time = random.uniform(0.5, 2.1)
            
            # Generate realistic demo result
            species_options = [
                "Wolf", "Deer", "Bear", "Fox", "Elk", "Moose", 
                "Cougar", "Coyote", "No_Animal", "Unsorted"
            ]
            
            # Weight the results to be more realistic
            weights = [0.08, 0.25, 0.12, 0.15, 0.18, 0.05, 0.03, 0.08, 0.04, 0.02]
            classification = random.choices(species_options, weights=weights)[0]
            
            # Generate confidence based on classification
            if classification == "No_Animal":
                confidence = random.uniform(0.1, 0.4)
            elif classification == "Unsorted":
                confidence = random.uniform(0.35, 0.65)
            else:
                confidence = random.uniform(0.65, 0.95)
            
            result = {
                'filename': filename,
                'classification': classification,
                'confidence': confidence,
                'processing_time': processing_time,
                'file_type': file_type,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            results.append(result)
            self.file_processed.emit(result)
        
        # Simulate final cleanup
        self.progress_updated.emit(98, "Generating reports...")
        time.sleep(0.3)
        
        self.progress_updated.emit(100, "Processing complete!")
        self.processing_complete.emit(results)
    
    def cancel(self):
        """Cancel the processing operation."""
        self.is_cancelled = True
        self.terminate()

class AdvancedSettingsDialog(QDialog):
    """
    Advanced settings configuration dialog with tabbed interface.
    
    This dialog allows users to fine-tune all processing parameters.
    To add new settings:
    1. Add the parameter to ProcessingSettings dataclass
    2. Add a control widget in the appropriate tab
    3. Update the accept() method to save the value
    """
    
    def __init__(self, settings: ProcessingSettings, parent=None):
        super().__init__(parent)
        self.settings = settings
        self.setupUI()
        self.connectSignals()
        
    def setupUI(self):
        """Setup the dialog user interface."""
        self.setWindowTitle("Advanced Processing Settings")
        self.setModal(True)
        self.resize(600, 700)
        
        # Apply dialog styling
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {THEME_CONFIG['background_color']};
            }}
        """)
        
        layout = QVBoxLayout(self)
        
        # Create tabs for different setting categories
        tabs = QTabWidget()
        
        # Video Processing Tab
        video_tab = self.create_video_settings_tab()
        tabs.addTab(video_tab, "Video Processing")
        
        # Image Processing Tab
        image_tab = self.create_image_settings_tab()
        tabs.addTab(image_tab, "Image Processing")
        
        # Advanced Tab
        advanced_tab = self.create_advanced_settings_tab()
        tabs.addTab(advanced_tab, "Advanced")
        
        # Export Tab
        export_tab = self.create_export_settings_tab()
        tabs.addTab(export_tab, "Export Options")
        
        layout.addWidget(tabs)
        
        # Dialog buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | 
            QDialogButtonBox.StandardButton.Cancel |
            QDialogButtonBox.StandardButton.RestoreDefaults
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        buttons.button(QDialogButtonBox.StandardButton.RestoreDefaults).clicked.connect(self.restore_defaults)
        
        layout.addWidget(buttons)
    
    def create_video_settings_tab(self) -> QWidget:
        """Create the video processing settings tab."""
        tab = QWidget()
        layout = QFormLayout(tab)
        
        # Confidence threshold
        self.confidence_spin = QDoubleSpinBox()
        self.confidence_spin.setRange(VALIDATION_RULES['min_confidence'], VALIDATION_RULES['max_confidence'])
        self.confidence_spin.setSingleStep(0.05)
        self.confidence_spin.setDecimals(3)
        self.confidence_spin.setValue(self.settings.confidence_threshold)
        self.confidence_spin.setSuffix(" (0.0-1.0)")
        layout.addRow("Confidence Threshold:", self.confidence_spin)
        
        # Dominant species threshold
        self.dominant_spin = QDoubleSpinBox()
        self.dominant_spin.setRange(0.5, 1.0)
        self.dominant_spin.setSingleStep(0.05)
        self.dominant_spin.setDecimals(3)
        self.dominant_spin.setValue(self.settings.dominant_species_threshold)
        self.dominant_spin.setSuffix(" (0.5-1.0)")
        layout.addRow("Dominant Species Threshold:", self.dominant_spin)
        
        # Max species transitions
        self.transitions_spin = QSpinBox()
        self.transitions_spin.setRange(1, 50)
        self.transitions_spin.setValue(self.settings.max_species_transitions)
        self.transitions_spin.setSuffix(" transitions")
        layout.addRow("Max Species Transitions:", self.transitions_spin)
        
        # Consecutive empty frames
        self.empty_frames_spin = QSpinBox()
        self.empty_frames_spin.setRange(1, 100)
        self.empty_frames_spin.setValue(self.settings.consecutive_empty_frames)
        self.empty_frames_spin.setSuffix(" frames")
        layout.addRow("Consecutive Empty Frames:", self.empty_frames_spin)
        
        # Add help text
        help_label = QLabel(
            "Video settings control how motion is analyzed across time. "
            "Higher thresholds make classification more strict."
        )
        help_label.setWordWrap(True)
        help_label.setStyleSheet(f"color: {THEME_CONFIG['text_secondary']}; font-style: italic;")
        layout.addRow(help_label)
        
        return tab
    
    def create_image_settings_tab(self) -> QWidget:
        """Create the image processing settings tab."""
        tab = QWidget()
        layout = QFormLayout(tab)
        
        # Image confidence threshold
        self.image_confidence_spin = QDoubleSpinBox()
        self.image_confidence_spin.setRange(VALIDATION_RULES['min_confidence'], VALIDATION_RULES['max_confidence'])
        self.image_confidence_spin.setSingleStep(0.05)
        self.image_confidence_spin.setDecimals(3)
        self.image_confidence_spin.setValue(self.settings.image_confidence_threshold)
        self.image_confidence_spin.setSuffix(" (0.0-1.0)")
        layout.addRow("Image Confidence Threshold:", self.image_confidence_spin)
        
        # Minimum detections
        self.min_detections_spin = QSpinBox()
        self.min_detections_spin.setRange(VALIDATION_RULES['min_detections'], VALIDATION_RULES['max_detections'])
        self.min_detections_spin.setValue(self.settings.image_min_detections)
        self.min_detections_spin.setSuffix(" detection(s)")
        layout.addRow("Minimum Detections Required:", self.min_detections_spin)
        
        # Multi-species threshold
        self.multi_species_spin = QDoubleSpinBox()
        self.multi_species_spin.setRange(0.0, 1.0)
        self.multi_species_spin.setSingleStep(0.05)
        self.multi_species_spin.setDecimals(3)
        self.multi_species_spin.setValue(self.settings.image_multi_species_threshold)
        self.multi_species_spin.setSuffix(" confidence diff")
        layout.addRow("Multi-Species Threshold:", self.multi_species_spin)
        
        # Unsorted confidence range
        unsorted_group = QGroupBox("Unsorted Classification Range")
        unsorted_layout = QFormLayout(unsorted_group)
        
        self.unsorted_min_spin = QDoubleSpinBox()
        self.unsorted_min_spin.setRange(0.0, 1.0)
        self.unsorted_min_spin.setSingleStep(0.05)
        self.unsorted_min_spin.setDecimals(3)
        self.unsorted_min_spin.setValue(self.settings.image_unsorted_min_confidence)
        unsorted_layout.addRow("Minimum:", self.unsorted_min_spin)
        
        self.unsorted_max_spin = QDoubleSpinBox()
        self.unsorted_max_spin.setRange(0.0, 1.0)
        self.unsorted_max_spin.setSingleStep(0.05)
        self.unsorted_max_spin.setDecimals(3)
        self.unsorted_max_spin.setValue(self.settings.image_unsorted_max_confidence)
        unsorted_layout.addRow("Maximum:", self.unsorted_max_spin)
        
        layout.addRow(unsorted_group)
        
        # Add help text
        help_label = QLabel(
            "Image settings apply to single-frame analysis. "
            "The unsorted range catches low-confidence detections for manual review."
        )
        help_label.setWordWrap(True)
        help_label.setStyleSheet(f"color: {THEME_CONFIG['text_secondary']}; font-style: italic;")
        layout.addRow(help_label)
        
        return tab
    
    def create_advanced_settings_tab(self) -> QWidget:
        """Create the advanced settings tab."""
        tab = QWidget()
        layout = QFormLayout(tab)
        
        # Processing options
        processing_group = QGroupBox("Processing Options")
        processing_layout = QFormLayout(processing_group)
        
        self.enable_gpu_checkbox = QCheckBox("Enable GPU Processing (if available)")
        self.enable_gpu_checkbox.setChecked(True)
        processing_layout.addRow(self.enable_gpu_checkbox)
        
        self.batch_size_spin = QSpinBox()
        self.batch_size_spin.setRange(1, 32)
        self.batch_size_spin.setValue(1)
        self.batch_size_spin.setSuffix(" images/batch")
        processing_layout.addRow("Batch Size:", self.batch_size_spin)
        
        self.max_threads_spin = QSpinBox()
        self.max_threads_spin.setRange(1, 16)
        self.max_threads_spin.setValue(4)
        self.max_threads_spin.setSuffix(" threads")
        processing_layout.addRow("Max Threads:", self.max_threads_spin)
        
        layout.addRow(processing_group)
        
        # Output options
        output_group = QGroupBox("Output Options")
        output_layout = QFormLayout(output_group)
        
        self.save_annotations_checkbox = QCheckBox("Save detection annotations")
        output_layout.addRow(self.save_annotations_checkbox)
        
        self.save_thumbnails_checkbox = QCheckBox("Save detection thumbnails")
        output_layout.addRow(self.save_thumbnails_checkbox)
        
        self.generate_summary_checkbox = QCheckBox("Generate processing summary")
        self.generate_summary_checkbox.setChecked(True)
        output_layout.addRow(self.generate_summary_checkbox)
        
        layout.addRow(output_group)
        
        return tab
    
    def create_export_settings_tab(self) -> QWidget:
        """Create the export settings tab."""
        tab = QWidget()
        layout = QFormLayout(tab)
        
        # Default export format
        self.export_format_combo = QComboBox()
        for key, info in EXPORT_FORMATS.items():
            self.export_format_combo.addItem(info['name'], key)
        layout.addRow("Default Export Format:", self.export_format_combo)
        
        # Export options
        export_group = QGroupBox("Export Options")
        export_layout = QFormLayout(export_group)
        
        self.include_timestamps_checkbox = QCheckBox("Include timestamps in exports")
        self.include_timestamps_checkbox.setChecked(True)
        export_layout.addRow(self.include_timestamps_checkbox)
        
        self.include_confidence_checkbox = QCheckBox("Include confidence scores")
        self.include_confidence_checkbox.setChecked(True)
        export_layout.addRow(self.include_confidence_checkbox)
        
        self.auto_open_checkbox = QCheckBox("Auto-open exported files")
        export_layout.addRow(self.auto_open_checkbox)
        
        layout.addRow(export_group)
        
        return tab
    
    def connectSignals(self):
        """Connect signals for real-time validation."""
        # Validate unsorted range when values change
        self.unsorted_min_spin.valueChanged.connect(self.validate_unsorted_range)
        self.unsorted_max_spin.valueChanged.connect(self.validate_unsorted_range)
    
    def validate_unsorted_range(self):
        """Ensure unsorted min is less than max."""
        if self.unsorted_min_spin.value() >= self.unsorted_max_spin.value():
            # Auto-adjust the other value
            if self.sender() == self.unsorted_min_spin:
                self.unsorted_max_spin.setValue(self.unsorted_min_spin.value() + 0.05)
            else:
                self.unsorted_min_spin.setValue(self.unsorted_max_spin.value() - 0.05)
    
    def restore_defaults(self):
        """Restore all settings to their default values."""
        reply = QMessageBox.question(
            self, "Restore Defaults",
            "This will reset all settings to their default values. Continue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # Reset to defaults
            defaults = DEFAULT_PROCESSING_SETTINGS
            self.confidence_spin.setValue(defaults['confidence_threshold'])
            self.dominant_spin.setValue(defaults['dominant_species_threshold'])
            self.transitions_spin.setValue(defaults['max_species_transitions'])
            self.empty_frames_spin.setValue(defaults['consecutive_empty_frames'])
            self.image_confidence_spin.setValue(defaults['image_confidence_threshold'])
            self.min_detections_spin.setValue(defaults['image_min_detections'])
            self.multi_species_spin.setValue(defaults['image_multi_species_threshold'])
            self.unsorted_min_spin.setValue(defaults['image_unsorted_min_confidence'])
            self.unsorted_max_spin.setValue(defaults['image_unsorted_max_confidence'])
    
    def accept(self):
        """Save settings and close dialog."""
        # Update settings object with dialog values
        self.settings.confidence_threshold = self.confidence_spin.value()
        self.settings.dominant_species_threshold = self.dominant_spin.value()
        self.settings.max_species_transitions = self.transitions_spin.value()
        self.settings.consecutive_empty_frames = self.empty_frames_spin.value()
        self.settings.image_confidence_threshold = self.image_confidence_spin.value()
        self.settings.image_min_detections = self.min_detections_spin.value()
        self.settings.image_multi_species_threshold = self.multi_species_spin.value()
        self.settings.image_unsorted_min_confidence = self.unsorted_min_spin.value()
        self.settings.image_unsorted_max_confidence = self.unsorted_max_spin.value()
        
        super().accept()

class ResultsWidget(QWidget):
    """
    Widget for displaying processing results with export capabilities.
    
    This widget shows a summary and detailed table of all processing results.
    It includes functionality to export data in various formats and copy to clipboard.
    """
    
    def __init__(self):
        super().__init__()
        self.results = []
        self.export_manager = ExportManager()
        self.setupUI()
        self.setupContextMenu()
        
    def setupUI(self):
        """Setup the results display interface."""
        layout = QVBoxLayout(self)
        
        # Summary section with statistics
        summary_group = QGroupBox("Processing Summary")
        summary_layout = QGridLayout(summary_group)
        
        # Create summary labels with better styling
        label_style = f"font-weight: bold; color: {THEME_CONFIG['text_primary']}; font-size: 16px;"
        value_style = f"color: {THEME_CONFIG['primary_color']}; font-size: 16px; font-weight: bold;"
        
        summary_layout.addWidget(QLabel("Total Files:"), 0, 0)
        self.total_files_label = QLabel("0")
        self.total_files_label.setStyleSheet(value_style)
        summary_layout.addWidget(self.total_files_label, 0, 1)
        
        summary_layout.addWidget(QLabel("Videos:"), 0, 2)
        self.videos_label = QLabel("0")
        self.videos_label.setStyleSheet(value_style)
        summary_layout.addWidget(self.videos_label, 0, 3)
        
        summary_layout.addWidget(QLabel("Images:"), 1, 0)
        self.images_label = QLabel("0")
        self.images_label.setStyleSheet(value_style)
        summary_layout.addWidget(self.images_label, 1, 1)
        
        summary_layout.addWidget(QLabel("Success Rate:"), 1, 2)
        self.success_rate_label = QLabel("0%")
        self.success_rate_label.setStyleSheet(value_style)
        summary_layout.addWidget(self.success_rate_label, 1, 3)
        
        layout.addWidget(summary_group)
        
        # Export controls
        export_layout = QHBoxLayout()
        
        export_layout.addWidget(QLabel("Export:"))
        
        self.copy_clipboard_btn = ModernButton("Copy to Clipboard", button_type="secondary")
        self.copy_clipboard_btn.clicked.connect(self.copy_to_clipboard)
        self.copy_clipboard_btn.setToolTip("Copy results in tab-separated format for Excel")
        export_layout.addWidget(self.copy_clipboard_btn)
        
        self.export_csv_btn = ModernButton("Export CSV", button_type="secondary")
        self.export_csv_btn.clicked.connect(lambda: self.export_data('csv'))
        export_layout.addWidget(self.export_csv_btn)
        
        self.export_excel_btn = ModernButton("Export Excel", button_type="secondary")
        self.export_excel_btn.clicked.connect(lambda: self.export_data('excel'))
        self.export_excel_btn.setEnabled(EXCEL_AVAILABLE)
        if not EXCEL_AVAILABLE:
            self.export_excel_btn.setToolTip("Install openpyxl for Excel export")
        export_layout.addWidget(self.export_excel_btn)
        
        self.export_custom_btn = ModernButton("Export...", button_type="secondary")
        self.export_custom_btn.clicked.connect(self.show_export_dialog)
        export_layout.addWidget(self.export_custom_btn)
        
        export_layout.addStretch()
        layout.addLayout(export_layout)
        
        # Results table with enhanced features
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([
            "Filename", "Classification", "Confidence", "Type", "Processing Time", "Timestamp"
        ])
        
        # Configure table appearance and behavior
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        
        # Add keyboard shortcuts
        copy_shortcut = QShortcut(QKeySequence.StandardKey.Copy, self.table)
        copy_shortcut.activated.connect(self.copy_selected_rows)
        
        layout.addWidget(self.table)
        
    def setupContextMenu(self):
        """Setup right-click context menu for the table."""
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        
    def show_context_menu(self, position):
        """Show context menu at the specified position."""
        if not self.table.itemAt(position):
            return
            
        menu = QMenu(self)
        
        # Copy actions
        copy_action = QAction("Copy Selected Rows", self)
        copy_action.triggered.connect(self.copy_selected_rows)
        menu.addAction(copy_action)
        
        copy_all_action = QAction("Copy All Data", self)
        copy_all_action.triggered.connect(self.copy_to_clipboard)
        menu.addAction(copy_all_action)
        
        menu.addSeparator()
        
        # Export actions
        export_csv_action = QAction("Export to CSV...", self)
        export_csv_action.triggered.connect(lambda: self.export_data('csv'))
        menu.addAction(export_csv_action)
        
        if EXCEL_AVAILABLE:
            export_excel_action = QAction("Export to Excel...", self)
            export_excel_action.triggered.connect(lambda: self.export_data('excel'))
            menu.addAction(export_excel_action)
        
        menu.addSeparator()
        
        # View actions
        expand_action = QAction("Expand All Columns", self)
        expand_action.triggered.connect(self.expand_columns)
        menu.addAction(expand_action)
        
        menu.exec(self.table.mapToGlobal(position))
        
    def expand_columns(self):
        """Expand all columns to fit content."""
        self.table.resizeColumnsToContents()
        
    def copy_selected_rows(self):
        """Copy selected table rows to clipboard."""
        selected_rows = set()
        for item in self.table.selectedItems():
            selected_rows.add(item.row())
            
        if not selected_rows:
            return
            
        # Get data for selected rows
        selected_data = []
        for row in sorted(selected_rows):
            if row < len(self.results):
                selected_data.append(self.results[row])
                
        self.export_manager.copy_to_clipboard(selected_data, 'tab_separated')
        
        # Show confirmation
        self.show_temporary_message(f"Copied {len(selected_data)} rows to clipboard")
        
    def add_result(self, result: dict):
        """
        Add a single result to the display.
        
        Args:
            result: Dictionary containing processing result data
        """
        self.results.append(result)
        self.update_display()
        
        # Auto-scroll to show new result
        self.table.scrollToBottom()
        
    def set_results(self, results: List[dict]):
        """
        Set all results at once.
        
        Args:
            results: List of result dictionaries
        """
        self.results = results
        self.update_display()
        
    def update_display(self):
        """Update the results display with current data."""
        # Update summary statistics
        total_files = len(self.results)
        videos = sum(1 for r in self.results if r.get('file_type') == 'video')
        images = total_files - videos
        
        # Calculate success rate (non-error results)
        successful = sum(1 for r in self.results if r.get('classification') not in ['Error', 'Failed'])
        success_rate = (successful / max(total_files, 1)) * 100
        
        self.total_files_label.setText(str(total_files))
        self.videos_label.setText(str(videos))
        self.images_label.setText(str(images))
        self.success_rate_label.setText(f"{success_rate:.1f}%")
        
        # Update table
        self.table.setRowCount(len(self.results))
        
        for row, result in enumerate(self.results):
            # Filename
            filename_item = QTableWidgetItem(result.get('filename', ''))
            filename_item.setToolTip(result.get('filename', ''))
            self.table.setItem(row, 0, filename_item)
            
            # Classification with color coding
            classification = result.get('classification', '')
            classification_item = QTableWidgetItem(classification)
            
            # Color-code classifications
            if classification == 'No_Animal':
                classification_item.setBackground(QColor(THEME_CONFIG['info_color']))
                classification_item.setForeground(QColor('white'))
            elif classification == 'Unsorted':
                classification_item.setBackground(QColor(THEME_CONFIG['warning_color']))
                classification_item.setForeground(QColor('black'))
            elif classification in ['Error', 'Failed']:
                classification_item.setBackground(QColor(THEME_CONFIG['error_color']))
                classification_item.setForeground(QColor('white'))
            else:
                classification_item.setBackground(QColor(THEME_CONFIG['success_color']))
                classification_item.setForeground(QColor('white'))
                
            self.table.setItem(row, 1, classification_item)
            
            # Confidence
            confidence = result.get('confidence', 0)
            confidence_item = QTableWidgetItem(f"{confidence:.3f}")
            confidence_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 2, confidence_item)
            
            # File type
            file_type_item = QTableWidgetItem(result.get('file_type', '').title())
            file_type_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 3, file_type_item)
            
            # Processing time
            processing_time = result.get('processing_time', 0)
            time_item = QTableWidgetItem(f"{processing_time:.2f}s")
            time_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 4, time_item)
            
            # Timestamp
            timestamp = result.get('timestamp', '')
            timestamp_item = QTableWidgetItem(timestamp)
            timestamp_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.table.setItem(row, 5, timestamp_item)
            
        # Enable/disable export buttons based on data availability
        has_data = len(self.results) > 0
        self.copy_clipboard_btn.setEnabled(has_data)
        self.export_csv_btn.setEnabled(has_data)
        self.export_excel_btn.setEnabled(has_data and EXCEL_AVAILABLE)
        self.export_custom_btn.setEnabled(has_data)
        
    def copy_to_clipboard(self):
        """Copy all results to clipboard in tab-separated format."""
        if not self.results:
            return
            
        success = self.export_manager.copy_to_clipboard(self.results, 'tab_separated')
        
        if success:
            self.show_temporary_message(f"Copied {len(self.results)} results to clipboard")
        else:
            self.show_temporary_message("Failed to copy to clipboard", is_error=True)
            
    def export_data(self, format_type: str):
        """
        Export data in the specified format.
        
        Args:
            format_type: Export format ('csv', 'excel', 'json', 'txt')
        """
        if not self.results:
            QMessageBox.information(self, "No Data", "No results to export.")
            return
            
        # Get file extension and filter
        format_info = EXPORT_FORMATS.get(format_type, {'name': 'Unknown', 'extension': '.txt'})
        extension = format_info['extension']
        filter_text = f"{format_info['name']} (*{extension});;All Files (*)"
        
        # Get save location
        default_filename = f"wolfvue_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}{extension}"
        file_path, _ = QFileDialog.getSaveFileName(
            self, f"Export {format_info['name']}", default_filename, filter_text
        )
        
        if not file_path:
            return
            
        # Perform export
        try:
            success = self.export_manager.export_data(self.results, file_path, format_type)
            
            if success:
                self.show_temporary_message(f"Successfully exported to {Path(file_path).name}")
                
                # Ask if user wants to open the file
                if DIALOG_SETTINGS['auto_close_success']:
                    reply = QMessageBox.question(
                        self, "Export Complete",
                        f"Data exported successfully to:\n{file_path}\n\nOpen the file now?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
                    )
                    
                    if reply == QMessageBox.StandardButton.Yes:
                        try:
                            import subprocess
                            if sys.platform.startswith('win'):
                                os.startfile(file_path)
                            elif sys.platform.startswith('darwin'):
                                subprocess.run(['open', file_path])
                            else:
                                subprocess.run(['xdg-open', file_path])
                        except Exception as e:
                            print(f"Could not open file: {e}")
            else:
                self.show_temporary_message("Export failed", is_error=True)
                
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export data:\n{str(e)}")
            
    def show_export_dialog(self):
        """Show custom export format dialog."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Custom Export")
        dialog.setModal(True)
        
        layout = QVBoxLayout(dialog)
        
        # Format selection
        format_group = QGroupBox("Export Format")
        format_layout = QVBoxLayout(format_group)
        
        format_combo = QComboBox()
        for key, info in EXPORT_FORMATS.items():
            format_combo.addItem(info['name'], key)
            
        format_layout.addWidget(format_combo)
        layout.addWidget(format_group)
        
        # Options
        options_group = QGroupBox("Options")
        options_layout = QVBoxLayout(options_group)
        
        include_summary_cb = QCheckBox("Include summary statistics")
        include_summary_cb.setChecked(True)
        options_layout.addWidget(include_summary_cb)
        
        include_metadata_cb = QCheckBox("Include processing metadata")
        include_metadata_cb.setChecked(True)
        options_layout.addWidget(include_metadata_cb)
        
        layout.addWidget(options_group)
        
        # Buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            format_type = format_combo.currentData()
            self.export_data(format_type)
            
    def show_temporary_message(self, message: str, is_error: bool = False):
        """
        Show a temporary status message.
        
        Args:
            message: Message to display
            is_error: Whether this is an error message
        """
        # Find the main window and show message in status bar
        main_window = self.window()
        if hasattr(main_window, 'statusBar'):
            if is_error:
                main_window.statusBar().setStyleSheet(f"color: {THEME_CONFIG['error_color']};")
            else:
                main_window.statusBar().setStyleSheet(f"color: {THEME_CONFIG['success_color']};")
                
            main_window.statusBar().showMessage(message, ANIMATION_CONFIG['status_display_time'])
            
            # Reset color after message expires
            QTimer.singleShot(
                ANIMATION_CONFIG['status_display_time'], 
                lambda: main_window.statusBar().setStyleSheet("")
            )

class WolfVueMainWindow(QMainWindow):
    """
    Main application window for WolfVue.
    
    This is the primary interface that users interact with. It coordinates
    all the other components and manages the overall application state.
    
    To customize the main window:
    1. Modify the LAYOUT_CONFIG and THEME_CONFIG at the top
    2. Override create_custom_widgets() to add new interface elements
    3. Add new menu items in setupMenus()
    4. Extend the settings in loadSettings()/saveSettings()
    """
    
    def __init__(self):
        super().__init__()
        
        # Initialize core components
        self.settings = ProcessingSettings()
        self.processing_worker = None
        self.theme_manager = ThemeManager()
        
        # Setup the interface
        self.setupUI()
        self.setupMenus()
        self.setupStatusBar()
        self.setupKeyboardShortcuts()
        
        # Load saved settings
        self.loadSettings()
        
        # Apply theme
        self.setStyleSheet(self.theme_manager.apply_theme())
        
    def setupUI(self):
        """Setup the main user interface layout and components."""
        self.setWindowTitle("WolfVue: Wildlife Video Classifier v1.0.0 Beta")
        self.setMinimumSize(LAYOUT_CONFIG['window_min_width'], LAYOUT_CONFIG['window_min_height'])
        
        # Set window icon if available
        try:
            icon_path = Path(__file__).parent / "resources" / "wolfvue_icon.png"
            if icon_path.exists():
                self.setWindowIcon(QIcon(str(icon_path)))
        except:
            pass  # Icon not available
        
        # Create central widget and main layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Main layout with splitter for resizable panels
        main_layout = QHBoxLayout(central_widget)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Left panel for configuration and controls
        config_panel = self.create_config_panel()
        splitter.addWidget(config_panel)
        
        # Right panel for results and progress
        results_panel = self.create_results_panel()
        splitter.addWidget(results_panel)
        
        # Set splitter proportions from configuration
        splitter.setSizes(LAYOUT_CONFIG['splitter_ratio'])
        splitter.setHandleWidth(2)
        
        main_layout.addWidget(splitter)
        
        # Add any custom widgets
        self.create_custom_widgets()
        
    def create_config_panel(self) -> QWidget:
        """Create the left configuration panel."""
        panel = QWidget()
        panel.setMaximumWidth(LAYOUT_CONFIG['left_panel_max_width'])
        layout = QVBoxLayout(panel)
        layout.setSpacing(LAYOUT_CONFIG['spacing_medium'])
        
        # Logo/Title section with gradient background
        title_widget = self.create_title_widget()
        layout.addWidget(title_widget)
        
        # File selection group
        file_group = QGroupBox("File Selection")
        file_layout = QVBoxLayout(file_group)
        
        # Input folder selection
        input_layout = QHBoxLayout()
        input_layout.addWidget(QLabel("Input Folder:"))
        
        self.input_folder_edit = QLineEdit()
        self.input_folder_edit.setPlaceholderText("Select folder containing videos/images...")
        input_layout.addWidget(self.input_folder_edit)
        
        input_browse_btn = ModernButton("Browse")
        input_browse_btn.clicked.connect(self.browse_input_folder)
        input_layout.addWidget(input_browse_btn)
        
        file_layout.addLayout(input_layout)
        
        # Output folder selection
        output_layout = QHBoxLayout()
        output_layout.addWidget(QLabel("Output Folder:"))
        
        self.output_folder_edit = QLineEdit()
        self.output_folder_edit.setPlaceholderText("Select output folder for sorted files...")
        output_layout.addWidget(self.output_folder_edit)
        
        output_browse_btn = ModernButton("Browse")
        output_browse_btn.clicked.connect(self.browse_output_folder)
        output_layout.addWidget(output_browse_btn)
        
        file_layout.addLayout(output_layout)
        
        layout.addWidget(file_group)
        
        # Model configuration group
        model_group = QGroupBox("Model Configuration")
        model_layout = QVBoxLayout(model_group)
        
        # Model file selection
        model_layout_h = QHBoxLayout()
        model_layout_h.addWidget(QLabel("Model File:"))
        
        self.model_path_edit = QLineEdit()
        self.model_path_edit.setPlaceholderText("Select YOLO model file (.pt)...")
        model_layout_h.addWidget(self.model_path_edit)
        
        model_browse_btn = ModernButton("Browse")
        model_browse_btn.clicked.connect(self.browse_model_file)
        model_layout_h.addWidget(model_browse_btn)
        
        model_layout.addLayout(model_layout_h)
        
        # Config file selection
        config_layout_h = QHBoxLayout()
        config_layout_h.addWidget(QLabel("Config File:"))
        
        self.config_file_edit = QLineEdit()
        self.config_file_edit.setPlaceholderText("Select configuration YAML file...")
        config_layout_h.addWidget(self.config_file_edit)
        
        config_browse_btn = ModernButton("Browse")
        config_browse_btn.clicked.connect(self.browse_config_file)
        config_layout_h.addWidget(config_browse_btn)
        
        model_layout.addLayout(config_layout_h)
        
        layout.addWidget(model_group)
        
        # Quick settings group
        quick_settings_group = QGroupBox("Quick Settings")
        quick_layout = QFormLayout(quick_settings_group)
        
        # Confidence threshold slider with live preview
        self.confidence_slider = QSlider(Qt.Orientation.Horizontal)
        self.confidence_slider.setRange(10, 95)
        self.confidence_slider.setValue(int(DEFAULT_PROCESSING_SETTINGS['confidence_threshold'] * 100))
        
        self.confidence_label = QLabel(f"{DEFAULT_PROCESSING_SETTINGS['confidence_threshold']:.2f}")
        self.confidence_label.setMinimumWidth(50)
        
        confidence_h_layout = QHBoxLayout()
        confidence_h_layout.addWidget(self.confidence_slider)
        confidence_h_layout.addWidget(self.confidence_label)
        
        # Connect slider to label update
        self.confidence_slider.valueChanged.connect(
            lambda v: self.confidence_label.setText(f"{v/100:.2f}")
        )
        
        quick_layout.addRow("Confidence Threshold:", confidence_h_layout)
        
        # Processing mode selection
        self.processing_mode_combo = QComboBox()
        self.processing_mode_combo.addItems(["Balanced", "High Precision", "High Recall", "Fast Processing"])
        self.processing_mode_combo.setCurrentText("Balanced")
        quick_layout.addRow("Processing Mode:", self.processing_mode_combo)
        
        layout.addWidget(quick_settings_group)
        
        # Control buttons section
        button_layout = QVBoxLayout()
        button_layout.setSpacing(LAYOUT_CONFIG['spacing_small'])
        
        # Main processing button
        self.process_btn = ModernButton("Start Processing", primary=True)
        self.process_btn.clicked.connect(self.start_processing)
        self.process_btn.setMinimumHeight(45)
        button_layout.addWidget(self.process_btn)
        
        # Stop button
        self.stop_btn = ModernButton("Stop Processing", button_type="danger")
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self.stop_processing)
        button_layout.addWidget(self.stop_btn)
        
        # Settings button
        settings_btn = ModernButton("Advanced Settings")
        settings_btn.clicked.connect(self.show_advanced_settings)
        button_layout.addWidget(settings_btn)
        
        # Theme button
        theme_btn = ModernButton("Customize Theme")
        theme_btn.clicked.connect(self.show_theme_settings)
        button_layout.addWidget(theme_btn)
        
        layout.addLayout(button_layout)
        layout.addStretch()
        
        return panel
        
    def create_title_widget(self) -> QWidget:
        """Create the title widget with gradient background."""
        title_widget = QWidget()
        title_layout = QVBoxLayout(title_widget)
        title_layout.setContentsMargins(LAYOUT_CONFIG['padding_large'], 
                                       LAYOUT_CONFIG['padding_large'],
                                       LAYOUT_CONFIG['padding_large'], 
                                       LAYOUT_CONFIG['padding_large'])
        
        # Main title
        title_label = QLabel("WolfVue")
        title_label.setStyleSheet(f"""
            QLabel {{
                font-size: {FONT_CONFIG['title_font_size']}px;
                font-weight: bold;
                color: white;
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {THEME_CONFIG['primary_color']}, stop:1 {THEME_CONFIG['primary_dark']});
                border-radius: {LAYOUT_CONFIG['border_radius'] + 4}px;
                padding: {LAYOUT_CONFIG['padding_large']}px;
            }}
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_layout.addWidget(title_label)
        
        # Subtitle
        subtitle_label = QLabel("Wildlife Video Classifier")
        subtitle_label.setStyleSheet(f"""
            QLabel {{
                font-size: {FONT_CONFIG['subtitle_font_size']}px;
                color: {THEME_CONFIG['text_secondary']};
                padding: 0px {LAYOUT_CONFIG['padding_large']}px {LAYOUT_CONFIG['padding_large']}px {LAYOUT_CONFIG['padding_large']}px;
            }}
        """)
        subtitle_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_layout.addWidget(subtitle_label)
        
        # Version and attribution
        version_label = QLabel("v1.0.0 Beta | Created by Nathan Bluto")
        version_label.setStyleSheet(f"""
            QLabel {{
                font-size: {FONT_CONFIG['small_font_size']}px;
                color: {THEME_CONFIG['text_muted']};
                font-style: italic;
            }}
        """)
        version_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_layout.addWidget(version_label)
        
        return title_widget
        
    def create_results_panel(self) -> QWidget:
        """Create the right results and progress panel."""
        panel = QWidget()
        layout = QVBoxLayout(panel)
        
        # Progress section
        progress_group = QGroupBox("Processing Progress")
        progress_layout = QVBoxLayout(progress_group)
        
        # Progress bar with custom styling
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setMinimumHeight(30)
        progress_layout.addWidget(self.progress_bar)
        
        # Status label with icon support
        self.status_label = QLabel("Ready to process files...")
        self.status_label.setStyleSheet(f"""
            QLabel {{ 
                color: {THEME_CONFIG['text_secondary']}; 
                font-style: italic; 
                padding: {LAYOUT_CONFIG['padding_small']}px;
            }}
        """)
        progress_layout.addWidget(self.status_label)
        
        # Processing statistics (shown during processing)
        stats_layout = QHBoxLayout()
        
        self.files_processed_label = QLabel("Files: 0/0")
        self.processing_speed_label = QLabel("Speed: 0 fps")
        self.eta_label = QLabel("ETA: --:--")
        
        stats_layout.addWidget(self.files_processed_label)
        stats_layout.addStretch()
        stats_layout.addWidget(self.processing_speed_label)
        stats_layout.addStretch()
        stats_layout.addWidget(self.eta_label)
        
        progress_layout.addLayout(stats_layout)
        
        layout.addWidget(progress_group)
        
        # Results section
        self.results_widget = ResultsWidget()
        layout.addWidget(self.results_widget)
        
        return panel
    
    def create_custom_widgets(self):
        """
        Override this method to add custom widgets to the interface.
        
        This is called after the main UI is set up, allowing for easy
        customization without modifying the core setup code.
        """
        pass  # Override in subclasses for custom functionality
        
    def setupMenus(self):
        """Setup application menus and actions."""
        menubar = self.menuBar()
        
        # File menu
        file_menu = menubar.addMenu("&File")
        
        new_action = QAction("&New Project", self)
        new_action.setShortcut(QKeySequence.StandardKey.New)
        new_action.triggered.connect(self.new_project)
        file_menu.addAction(new_action)
        
        open_action = QAction("&Open Project...", self)
        open_action.setShortcut(QKeySequence.StandardKey.Open)
        open_action.triggered.connect(self.open_project)
        file_menu.addAction(open_action)
        
        save_action = QAction("&Save Project", self)
        save_action.setShortcut(QKeySequence.StandardKey.Save)
        save_action.triggered.connect(self.save_project)
        file_menu.addAction(save_action)
        
        file_menu.addSeparator()
        
        export_menu = file_menu.addMenu("&Export Results")
        
        export_csv_action = QAction("Export as &CSV...", self)
        export_csv_action.triggered.connect(lambda: self.results_widget.export_data('csv'))
        export_menu.addAction(export_csv_action)
        
        if EXCEL_AVAILABLE:
            export_excel_action = QAction("Export as &Excel...", self)
            export_excel_action.triggered.connect(lambda: self.results_widget.export_data('excel'))
            export_menu.addAction(export_excel_action)
        
        file_menu.addSeparator()
        
        exit_action = QAction("E&xit", self)
        exit_action.setShortcut(QKeySequence.StandardKey.Quit)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)
        
        # Edit menu
        edit_menu = menubar.addMenu("&Edit")
        
        copy_action = QAction("&Copy Results", self)
        copy_action.setShortcut(QKeySequence.StandardKey.Copy)
        copy_action.triggered.connect(self.results_widget.copy_to_clipboard)
        edit_menu.addAction(copy_action)
        
        edit_menu.addSeparator()
        
        settings_action = QAction("&Preferences...", self)
        settings_action.setShortcut(QKeySequence.StandardKey.Preferences)
        settings_action.triggered.connect(self.show_advanced_settings)
        edit_menu.addAction(settings_action)
        
        # View menu
        view_menu = menubar.addMenu("&View")
        
        theme_action = QAction("&Theme Settings...", self)
        theme_action.triggered.connect(self.show_theme_settings)
        view_menu.addAction(theme_action)
        
        view_menu.addSeparator()
        
        fullscreen_action = QAction("&Full Screen", self)
        fullscreen_action.setShortcut(QKeySequence.StandardKey.FullScreen)
        fullscreen_action.triggered.connect(self.toggle_fullscreen)
        view_menu.addAction(fullscreen_action)
        
        # Tools menu
        tools_menu = menubar.addMenu("&Tools")
        
        advanced_settings_action = QAction("&Advanced Settings...", self)
        advanced_settings_action.triggered.connect(self.show_advanced_settings)
        tools_menu.addAction(advanced_settings_action)
        
        batch_process_action = QAction("&Batch Process Multiple Folders...", self)
        batch_process_action.triggered.connect(self.show_batch_dialog)
        tools_menu.addAction(batch_process_action)
        
        # Help menu
        help_menu = menubar.addMenu("&Help")
        
        docs_action = QAction("&Documentation", self)
        docs_action.triggered.connect(self.show_documentation)
        help_menu.addAction(docs_action)
        
        help_menu.addSeparator()
        
        about_action = QAction("&About WolfVue", self)
        about_action.triggered.connect(self.show_about)
        help_menu.addAction(about_action)
        
    def setupStatusBar(self):
        """Setup the status bar with additional information."""
        status_bar = self.statusBar()
        status_bar.showMessage("Ready")
        
        # Add permanent widgets to status bar
        self.yolo_status_label = QLabel("YOLO: Available" if YOLO_AVAILABLE else "YOLO: Demo Mode")
        self.yolo_status_label.setStyleSheet(f"""
            QLabel {{
                color: {THEME_CONFIG['success_color'] if YOLO_AVAILABLE else THEME_CONFIG['warning_color']};
                font-weight: bold;
                padding: 0 10px;
            }}
        """)
        status_bar.addPermanentWidget(self.yolo_status_label)
        
    def setupKeyboardShortcuts(self):
        """Setup keyboard shortcuts for common actions."""
        # Processing shortcuts
        start_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
        start_shortcut.activated.connect(self.start_processing)
        
        stop_shortcut = QShortcut(QKeySequence("Ctrl+S"), self)
        stop_shortcut.activated.connect(self.stop_processing)
        
        # Settings shortcut
        settings_shortcut = QShortcut(QKeySequence("Ctrl+,"), self)
        settings_shortcut.activated.connect(self.show_advanced_settings)
        
    # ========================================================================
    # FILE AND FOLDER SELECTION METHODS
    # ========================================================================
    
    def browse_input_folder(self):
        """Browse for input folder containing videos and images."""
        folder = QFileDialog.getExistingDirectory(
            self, "Select Input Folder", self.settings.input_folder,
            QFileDialog.Option.ShowDirsOnly | QFileDialog.Option.DontResolveSymlinks
        )
        if folder:
            self.input_folder_edit.setText(folder)
            self.settings.input_folder = folder
            self.validate_inputs()
            
    def browse_output_folder(self):
        """Browse for output folder where sorted files will be saved."""
        folder = QFileDialog.getExistingDirectory(
            self, "Select Output Folder", self.settings.output_folder,
            QFileDialog.Option.ShowDirsOnly | QFileDialog.Option.DontResolveSymlinks
        )
        if folder:
            self.output_folder_edit.setText(folder)
            self.settings.output_folder = folder
            self.validate_inputs()
            
    def browse_model_file(self):
        """Browse for YOLO model file."""
        model_filter = "Model Files ("
        for ext in SUPPORTED_FILE_TYPES['model_extensions']:
            model_filter += f"*{ext} "
        model_filter += ");;All Files (*)"
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select YOLO Model File", self.settings.model_path, model_filter
        )
        if file_path:
            self.model_path_edit.setText(file_path)
            self.settings.model_path = file_path
            self.validate_inputs()
            
    def browse_config_file(self):
        """Browse for configuration YAML file."""
        config_filter = "Config Files ("
        for ext in SUPPORTED_FILE_TYPES['config_extensions']:
            config_filter += f"*{ext} "
        config_filter += ");;All Files (*)"
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Select Configuration File", self.settings.config_file, config_filter
        )
        if file_path:
            self.config_file_edit.setText(file_path)
            self.settings.config_file = file_path
            self.validate_inputs()
    
    def validate_inputs(self):
        """Validate user inputs and enable/disable processing button."""
        has_input = bool(self.settings.input_folder and os.path.exists(self.settings.input_folder))
        has_output = bool(self.settings.output_folder)
        
        # Enable processing if we have minimum required inputs
        can_process = has_input and has_output
        self.process_btn.setEnabled(can_process and not self.is_processing())
        
        # Update status message based on validation
        if not has_input:
            self.status_label.setText("Please select an input folder...")
        elif not has_output:
            self.status_label.setText("Please select an output folder...")
        elif not YOLO_AVAILABLE and not self.settings.model_path:
            self.status_label.setText("YOLO not available - will run in demo mode...")
        else:
            self.status_label.setText("Ready to process files...")
    
    # ========================================================================
    # DIALOG AND SETTINGS METHODS
    # ========================================================================
    
    def show_advanced_settings(self):
        """Show the advanced settings dialog."""
        dialog = AdvancedSettingsDialog(self.settings, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.saveSettings()
            self.statusBar().showMessage("Settings updated", 2000)
            
    def show_theme_settings(self):
        """Show theme customization dialog."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Theme Settings")
        dialog.setModal(True)
        dialog.resize(400, 300)
        
        layout = QVBoxLayout(dialog)
        
        # Color selection
        colors_group = QGroupBox("Theme Colors")
        colors_layout = QFormLayout(colors_group)
        
        # Primary color
        primary_color_btn = QPushButton()
        primary_color_btn.setStyleSheet(f"background-color: {THEME_CONFIG['primary_color']}; min-height: 30px;")
        primary_color_btn.clicked.connect(lambda: self.choose_color('primary_color', primary_color_btn))
        colors_layout.addRow("Primary Color:", primary_color_btn)
        
        # Background color
        bg_color_btn = QPushButton()
        bg_color_btn.setStyleSheet(f"background-color: {THEME_CONFIG['background_color']}; min-height: 30px;")
        bg_color_btn.clicked.connect(lambda: self.choose_color('background_color', bg_color_btn))
        colors_layout.addRow("Background Color:", bg_color_btn)
        
        layout.addWidget(colors_group)
        
        # Font selection
        font_group = QGroupBox("Font Settings")
        font_layout = QFormLayout(font_group)
        
        font_btn = QPushButton(f"{FONT_CONFIG['main_font_family']} - {FONT_CONFIG['main_font_size']}pt")
        font_btn.clicked.connect(lambda: self.choose_font(font_btn))
        font_layout.addRow("Main Font:", font_btn)
        
        layout.addWidget(font_group)
        
        # Preset themes
        presets_group = QGroupBox("Theme Presets")
        presets_layout = QVBoxLayout(presets_group)
        
        preset_buttons_layout = QHBoxLayout()
        
        dark_theme_btn = ModernButton("Dark Theme")
        dark_theme_btn.clicked.connect(lambda: self.apply_preset_theme('dark'))
        preset_buttons_layout.addWidget(dark_theme_btn)
        
        light_theme_btn = ModernButton("Light Theme")
        light_theme_btn.clicked.connect(lambda: self.apply_preset_theme('light'))
        preset_buttons_layout.addWidget(light_theme_btn)
        
        nature_theme_btn = ModernButton("Nature Theme")
        nature_theme_btn.clicked.connect(lambda: self.apply_preset_theme('nature'))
        preset_buttons_layout.addWidget(nature_theme_btn)
        
        presets_layout.addLayout(preset_buttons_layout)
        layout.addWidget(presets_group)
        
        # Dialog buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.setStyleSheet(self.theme_manager.apply_theme())
            
    def choose_color(self, color_key: str, button: QPushButton):
        """Choose a color for the theme."""
        current_color = QColor(THEME_CONFIG[color_key])
        color = QColorDialog.getColor(current_color, self, f"Choose {color_key.replace('_', ' ').title()}")
        
        if color.isValid():
            THEME_CONFIG[color_key] = color.name()
            button.setStyleSheet(f"background-color: {color.name()}; min-height: 30px;")
            
    def choose_font(self, button: QPushButton):
        """Choose a font for the interface."""
        current_font = QFont(FONT_CONFIG['main_font_family'], FONT_CONFIG['main_font_size'])
        font, ok = QFontDialog.getFont(current_font, self, "Choose Main Font")
        
        if ok:
            FONT_CONFIG['main_font_family'] = font.family()
            FONT_CONFIG['main_font_size'] = font.pointSize()
            button.setText(f"{font.family()} - {font.pointSize()}pt")
            
    def apply_preset_theme(self, theme_name: str):
        """Apply a preset theme configuration."""
        if theme_name == 'dark':
            THEME_CONFIG.update({
                'primary_color': '#BB86FC',
                'background_color': '#121212',
                'surface_color': '#1F1F1F',
                'border_color': '#333333',
                'text_primary': '#FFFFFF',
                'text_secondary': '#CCCCCC',
                'text_muted': '#888888',
            })
        elif theme_name == 'light':
            THEME_CONFIG.update({
                'primary_color': '#2196F3',
                'background_color': '#FAFAFA',
                'surface_color': '#FFFFFF',
                'border_color': '#E0E0E0',
                'text_primary': '#212121',
                'text_secondary': '#757575',
                'text_muted': '#BDBDBD',
            })
        elif theme_name == 'nature':
            THEME_CONFIG.update({
                'primary_color': '#4CAF50',
                'background_color': '#F1F8E9',
                'surface_color': '#FFFFFF',
                'border_color': '#C8E6C9',
                'text_primary': '#2E7D32',
                'text_secondary': '#388E3C',
                'text_muted': '#81C784',
            })
        
        self.setStyleSheet(self.theme_manager.apply_theme())
    
    # ========================================================================
    # PROCESSING CONTROL METHODS
    # ========================================================================
    
    def start_processing(self):
        """Start the video and image processing workflow."""
        # Validate inputs before starting
        if not self.validate_processing_inputs():
            return
            
        # Update settings from UI elements
        self.update_settings_from_ui()
        
        # Create output folder if it doesn't exist
        try:
            os.makedirs(self.settings.output_folder, exist_ok=True)
        except OSError as e:
            QMessageBox.critical(self, "Error", f"Could not create output folder:\n{str(e)}")
            return
        
        # Show confirmation dialog for large operations
        if self.should_show_processing_confirmation():
            if not self.show_processing_confirmation():
                return
        
        # Start the processing worker thread
        self.processing_worker = ProcessingWorker(self.settings)
        self.processing_worker.progress_updated.connect(self.update_progress)
        self.processing_worker.file_processed.connect(self.on_file_processed)
        self.processing_worker.processing_complete.connect(self.on_processing_complete)
        self.processing_worker.error_occurred.connect(self.on_processing_error)
        
        self.processing_worker.start()
        
        # Update UI state for processing
        self.set_processing_ui_state(True)
        
        # Clear previous results
        self.results_widget.set_results([])
        
        self.statusBar().showMessage("Processing started...")
        
    def stop_processing(self):
        """Stop the current processing operation."""
        if not self.processing_worker or not self.processing_worker.isRunning():
            return
            
        # Confirm stop action
        reply = QMessageBox.question(
            self, "Stop Processing",
            "Are you sure you want to stop processing?\nProgress will be lost.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.processing_worker.cancel()
            self.processing_worker.wait(3000)  # Wait up to 3 seconds
            
            self.set_processing_ui_state(False)
            self.progress_bar.setValue(0)
            self.status_label.setText("Processing stopped by user.")
            self.statusBar().showMessage("Ready")
    
    def validate_processing_inputs(self) -> bool:
        """Validate all inputs required for processing."""
        # Check input folder
        if not self.settings.input_folder or not os.path.exists(self.settings.input_folder):
            QMessageBox.warning(self, "Invalid Input", "Please select a valid input folder.")
            return False
            
        # Check output folder
        if not self.settings.output_folder:
            QMessageBox.warning(self, "Invalid Output", "Please select an output folder.")
            return False
            
        # Check if we have YOLO or demo mode
        if not YOLO_AVAILABLE:
            reply = QMessageBox.question(
                self, "Demo Mode", 
                "YOLO dependencies not available. Run in demo mode?\n\n"
                "Demo mode will simulate processing with fake results.",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return False
        
        # Check for model file if YOLO is available
        elif YOLO_AVAILABLE and self.settings.model_path and not os.path.exists(self.settings.model_path):
            reply = QMessageBox.question(
                self, "Model File Not Found",
                f"Model file not found: {self.settings.model_path}\n\n"
                "Continue with default model?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return False
        
        return True
    
    def update_settings_from_ui(self):
        """Update settings object from UI controls."""
        self.settings.input_folder = self.input_folder_edit.text().strip()
        self.settings.output_folder = self.output_folder_edit.text().strip()
        self.settings.model_path = self.model_path_edit.text().strip()
        self.settings.config_file = self.config_file_edit.text().strip()
        self.settings.confidence_threshold = self.confidence_slider.value() / 100
        
        # Apply processing mode adjustments
        mode = self.processing_mode_combo.currentText()
        if mode == "High Precision":
            self.settings.confidence_threshold = min(self.settings.confidence_threshold + 0.1, 0.95)
            self.settings.dominant_species_threshold = 0.95
        elif mode == "High Recall":
            self.settings.confidence_threshold = max(self.settings.confidence_threshold - 0.1, 0.1)
            self.settings.dominant_species_threshold = 0.7
        elif mode == "Fast Processing":
            self.settings.consecutive_empty_frames = 5
            
    def should_show_processing_confirmation(self) -> bool:
        """Determine if we should show a confirmation dialog before processing."""
        # Show confirmation for operations that might take a long time
        try:
            input_path = Path(self.settings.input_folder)
            file_count = 0
            
            for ext in SUPPORTED_FILE_TYPES['video_extensions'] + SUPPORTED_FILE_TYPES['image_extensions']:
                file_count += len(list(input_path.glob(f'*{ext}')))
                file_count += len(list(input_path.glob(f'*{ext.upper()}')))
                
            return file_count > 50  # Show confirmation for more than 50 files
        except:
            return False
    
    def show_processing_confirmation(self) -> bool:
        """Show processing confirmation dialog."""
        try:
            input_path = Path(self.settings.input_folder)
            file_count = 0
            
            for ext in SUPPORTED_FILE_TYPES['video_extensions'] + SUPPORTED_FILE_TYPES['image_extensions']:
                file_count += len(list(input_path.glob(f'*{ext}')))
                
            estimated_time = file_count * 2  # Rough estimate: 2 seconds per file
            
            message = (
                f"About to process {file_count} files.\n"
                f"Estimated processing time: {estimated_time // 60} minutes\n\n"
                f"Input: {self.settings.input_folder}\n"
                f"Output: {self.settings.output_folder}\n\n"
                f"Continue with processing?"
            )
            
            reply = QMessageBox.question(
                self, "Confirm Processing", message,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.Yes
            )
            
            return reply == QMessageBox.StandardButton.Yes
            
        except Exception:
            return True  # Default to allowing processing
    
    def set_processing_ui_state(self, processing: bool):
        """Update UI elements based on processing state."""
        # Enable/disable controls
        self.process_btn.setEnabled(not processing)
        self.stop_btn.setEnabled(processing)
        
        # Disable input controls during processing
        self.input_folder_edit.setEnabled(not processing)
        self.output_folder_edit.setEnabled(not processing)
        self.model_path_edit.setEnabled(not processing)
        self.config_file_edit.setEnabled(not processing)
        
        # Update progress bar visibility
        if processing:
            self.progress_bar.setValue(0)
            self.progress_bar.setVisible(True)
        else:
            self.progress_bar.setVisible(not processing)
    
    def is_processing(self) -> bool:
        """Check if processing is currently active."""
        return self.processing_worker and self.processing_worker.isRunning()
    
    # ========================================================================
    # PROCESSING EVENT HANDLERS
    # ========================================================================
    
    @pyqtSlot(int, str)
    def update_progress(self, progress: int, status: str):
        """Update progress display with current processing status."""
        self.progress_bar.setValue(progress)
        self.status_label.setText(status)
        
        # Update additional statistics if available
        if hasattr(self, 'start_time'):
            elapsed = time.time() - self.start_time
            if progress > 0:
                estimated_total = elapsed * 100 / progress
                remaining = estimated_total - elapsed
                self.eta_label.setText(f"ETA: {timedelta(seconds=int(remaining))}")
        
    @pyqtSlot(dict)
    def on_file_processed(self, result: dict):
        """Handle completion of a single file."""
        self.results_widget.add_result(result)
        
        # Update file counter
        current_count = len(self.results_widget.results)
        self.files_processed_label.setText(f"Files: {current_count}")
        
        # Log successful processing
        if result.get('classification') not in ['Error', 'Failed']:
            self.statusBar().showMessage(
                f"Processed: {result.get('filename', 'Unknown')} -> {result.get('classification', 'Unknown')}", 
                1000
            )
        
    @pyqtSlot(list)
    def on_processing_complete(self, results: List[dict]):
        """Handle completion of all processing."""
        self.set_processing_ui_state(False)
        
        # Update status
        total_files = len(results)
        successful = sum(1 for r in results if r.get('classification') not in ['Error', 'Failed'])
        
        completion_message = (
            f"Processing complete! Successfully processed {successful}/{total_files} files."
        )
        
        self.status_label.setText(completion_message)
        self.statusBar().showMessage("Processing complete")
        
        # Show completion dialog
        completion_dialog = QMessageBox(self)
        completion_dialog.setWindowTitle("Processing Complete")
        completion_dialog.setText(completion_message)
        completion_dialog.setDetailedText(
            f"Results saved to: {self.settings.output_folder}\n\n"
            f"Processing summary:\n"
            f"- Total files: {total_files}\n"
            f"- Successful: {successful}\n"
            f"- Failed: {total_files - successful}\n"
            f"- Success rate: {(successful/total_files*100):.1f}%"
        )
        completion_dialog.setIcon(QMessageBox.Icon.Information)
        completion_dialog.exec()
        
    @pyqtSlot(str)
    def on_processing_error(self, error_message: str):
        """Handle processing errors."""
        self.stop_processing()
        
        error_dialog = QMessageBox(self)
        error_dialog.setWindowTitle("Processing Error")
        error_dialog.setText("An error occurred during processing:")
        error_dialog.setDetailedText(error_message)
        error_dialog.setIcon(QMessageBox.Icon.Critical)
        error_dialog.exec()
        
    # ========================================================================
    # PROJECT MANAGEMENT METHODS
    # ========================================================================
    
    def new_project(self):
        """Start a new project, clearing all current settings."""
        if self.is_processing():
            QMessageBox.warning(self, "Processing Active", 
                              "Cannot start new project while processing is active.")
            return
            
        reply = QMessageBox.question(
            self, "New Project",
            "This will clear all current settings and results. Continue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # Reset settings to defaults
            self.settings = ProcessingSettings()
            
            # Clear UI
            self.input_folder_edit.clear()
            self.output_folder_edit.clear()
            self.model_path_edit.clear()
            self.config_file_edit.clear()
            self.confidence_slider.setValue(int(DEFAULT_PROCESSING_SETTINGS['confidence_threshold'] * 100))
            self.processing_mode_combo.setCurrentText("Balanced")
            
            # Clear results
            self.results_widget.set_results([])
            self.progress_bar.setValue(0)
            self.status_label.setText("Ready to process files...")
            
            self.statusBar().showMessage("New project created", 2000)
    
    def save_project(self):
        """Save the current project settings."""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Project", "", "WolfVue Projects (*.wvproj);;JSON Files (*.json)"
        )
        
        if file_path:
            try:
                project_data = {
                    'settings': asdict(self.settings),
                    'theme': THEME_CONFIG,
                    'font': FONT_CONFIG,
                    'version': '1.0.0',
                    'saved_at': datetime.now().isoformat()
                }
                
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(project_data, f, indent=2)
                    
                self.statusBar().showMessage(f"Project saved to {Path(file_path).name}", 3000)
                
            except Exception as e:
                QMessageBox.critical(self, "Save Error", f"Could not save project:\n{str(e)}")
    
    def open_project(self):
        """Open a saved project."""
        if self.is_processing():
            QMessageBox.warning(self, "Processing Active", 
                              "Cannot open project while processing is active.")
            return
            
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Project", "", "WolfVue Projects (*.wvproj);;JSON Files (*.json)"
        )
        
        if file_path:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    project_data = json.load(f)
                
                # Load settings
                if 'settings' in project_data:
                    settings_dict = project_data['settings']
                    self.settings = ProcessingSettings(**settings_dict)
                    
                    # Update UI with loaded settings
                    self.input_folder_edit.setText(self.settings.input_folder)
                    self.output_folder_edit.setText(self.settings.output_folder)
                    self.model_path_edit.setText(self.settings.model_path)
                    self.config_file_edit.setText(self.settings.config_file)
                    self.confidence_slider.setValue(int(self.settings.confidence_threshold * 100))
                
                # Load theme if present
                if 'theme' in project_data:
                    THEME_CONFIG.update(project_data['theme'])
                    self.setStyleSheet(self.theme_manager.apply_theme())
                
                self.statusBar().showMessage(f"Project loaded from {Path(file_path).name}", 3000)
                
            except Exception as e:
                QMessageBox.critical(self, "Load Error", f"Could not load project:\n{str(e)}")
    
    # ========================================================================
    # ADDITIONAL FEATURE METHODS
    # ========================================================================
    
    def show_batch_dialog(self):
        """Show batch processing dialog for multiple folders."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Batch Process Multiple Folders")
        dialog.setModal(True)
        dialog.resize(600, 400)
        
        layout = QVBoxLayout(dialog)
        
        # Instructions
        instructions = QLabel(
            "Select multiple input folders to process in batch. "
            "Each folder will be processed with the same settings."
        )
        instructions.setWordWrap(True)
        layout.addWidget(instructions)
        
        # Folder list
        folders_group = QGroupBox("Input Folders")
        folders_layout = QVBoxLayout(folders_group)
        
        folder_list = QListWidget()
        folders_layout.addWidget(folder_list)
        
        folder_buttons_layout = QHBoxLayout()
        
        add_folder_btn = ModernButton("Add Folder")
        add_folder_btn.clicked.connect(lambda: self.add_batch_folder(folder_list))
        folder_buttons_layout.addWidget(add_folder_btn)
        
        remove_folder_btn = ModernButton("Remove Selected")
        remove_folder_btn.clicked.connect(lambda: self.remove_batch_folder(folder_list))
        folder_buttons_layout.addWidget(remove_folder_btn)
        
        folder_buttons_layout.addStretch()
        folders_layout.addLayout(folder_buttons_layout)
        
        layout.addWidget(folders_group)
        
        # Output settings
        output_group = QGroupBox("Output Settings")
        output_layout = QFormLayout(output_group)
        
        base_output_edit = QLineEdit()
        base_output_browse = ModernButton("Browse")
        base_output_browse.clicked.connect(
            lambda: self.browse_batch_output(base_output_edit)
        )
        
        output_h_layout = QHBoxLayout()
        output_h_layout.addWidget(base_output_edit)
        output_h_layout.addWidget(base_output_browse)
        
        output_layout.addRow("Base Output Folder:", output_h_layout)
        
        create_subfolders_cb = QCheckBox("Create subfolders for each input")
        create_subfolders_cb.setChecked(True)
        output_layout.addRow(create_subfolders_cb)
        
        layout.addWidget(output_group)
        
        # Dialog buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Process batch folders
            folders = [folder_list.item(i).text() for i in range(folder_list.count())]
            base_output = base_output_edit.text()
            create_subs = create_subfolders_cb.isChecked()
            
            if folders and base_output:
                self.process_batch_folders(folders, base_output, create_subs)
    
    def add_batch_folder(self, folder_list: QListWidget):
        """Add a folder to the batch processing list."""
        folder = QFileDialog.getExistingDirectory(self, "Select Input Folder")
        if folder and folder not in [folder_list.item(i).text() for i in range(folder_list.count())]:
            folder_list.addItem(folder)
    
    def remove_batch_folder(self, folder_list: QListWidget):
        """Remove selected folder from batch processing list."""
        current_row = folder_list.currentRow()
        if current_row >= 0:
            folder_list.takeItem(current_row)
    
    def browse_batch_output(self, edit_widget: QLineEdit):
        """Browse for batch output folder."""
        folder = QFileDialog.getExistingDirectory(self, "Select Base Output Folder")
        if folder:
            edit_widget.setText(folder)
    
    def process_batch_folders(self, folders: List[str], base_output: str, create_subfolders: bool):
        """Process multiple folders in batch."""
        # This would implement batch processing logic
        QMessageBox.information(
            self, "Batch Processing", 
            f"Batch processing of {len(folders)} folders is not yet implemented in this demo."
        )
    
    def show_documentation(self):
        """Show documentation or help."""
        help_dialog = QDialog(self)
        help_dialog.setWindowTitle("WolfVue Documentation")
        help_dialog.setModal(True)
        help_dialog.resize(700, 500)
        
        layout = QVBoxLayout(help_dialog)
        
        # Create tabbed help content
        help_tabs = QTabWidget()
        
        # Quick Start tab
        quick_start_text = QTextEdit()
        quick_start_text.setReadOnly(True)
        quick_start_text.setHtml("""
        <h2>Quick Start Guide</h2>
        <h3>1. Select Input Folder</h3>
        <p>Choose the folder containing your videos and images to classify.</p>
        
        <h3>2. Select Output Folder</h3>
        <p>Choose where you want the sorted files to be saved.</p>
        
        <h3>3. Configure Model (Optional)</h3>
        <p>Select your YOLO model file and configuration. If not specified, default settings will be used.</p>
        
        <h3>4. Adjust Settings</h3>
        <p>Use the confidence slider to adjust detection sensitivity. Higher values = more strict.</p>
        
        <h3>5. Start Processing</h3>
        <p>Click "Start Processing" to begin classification.</p>
        
        <h3>6. Export Results</h3>
        <p>Use the export buttons to save results in CSV, Excel, or other formats.</p>
        """)
        help_tabs.addTab(quick_start_text, "Quick Start")
        
        # Settings tab
        settings_text = QTextEdit()
        settings_text.setReadOnly(True)
        settings_text.setHtml("""
        <h2>Settings Guide</h2>
        <h3>Confidence Threshold</h3>
        <p>Minimum confidence required for a detection to be considered valid. 
        Range: 0.0 to 1.0. Higher values = fewer false positives.</p>
        
        <h3>Processing Modes</h3>
        <ul>
        <li><b>Balanced:</b> Default settings for most use cases</li>
        <li><b>High Precision:</b> Reduces false positives, may miss some detections</li>
        <li><b>High Recall:</b> Catches more animals, may have more false positives</li>
        <li><b>Fast Processing:</b> Optimized for speed</li>
        </ul>
        
        <h3>Advanced Settings</h3>
        <p>Access through Tools > Advanced Settings for fine-tuning algorithm parameters.</p>
        """)
        help_tabs.addTab(settings_text, "Settings")
        
        # Export tab
        export_text = QTextEdit()
        export_text.setReadOnly(True)
        export_text.setHtml("""
        <h2>Export Options</h2>
        <h3>Copy to Clipboard</h3>
        <p>Copies results in tab-separated format, perfect for pasting into Excel.</p>
        
        <h3>CSV Export</h3>
        <p>Saves results as a comma-separated values file for use in any spreadsheet application.</p>
        
        <h3>Excel Export</h3>
        <p>Creates a formatted Excel file with summary statistics and detailed results.</p>
        
        <h3>Custom Export</h3>
        <p>Additional formats including JSON and formatted text reports.</p>
        
        <h3>Keyboard Shortcuts</h3>
        <ul>
        <li><b>Ctrl+C:</b> Copy results to clipboard</li>
        <li><b>Ctrl+R:</b> Start processing</li>
        <li><b>Ctrl+S:</b> Stop processing</li>
        <li><b>Ctrl+,:</b> Open settings</li>
        </ul>
        """)
        help_tabs.addTab(export_text, "Export & Shortcuts")
        
        layout.addWidget(help_tabs)
        
        # Close button
        close_btn = ModernButton("Close")
        close_btn.clicked.connect(help_dialog.accept)
        layout.addWidget(close_btn)
        
        help_dialog.exec()
    
    def toggle_fullscreen(self):
        """Toggle fullscreen mode."""
        if self.isFullScreen():
            self.showNormal()
        else:
            self.showFullScreen()
    
    def show_about(self):
        """Show about dialog with application information."""
        about_text = f"""
        <div style="text-align: center;">
        <h1 style="color: {THEME_CONFIG['primary_color']};">WolfVue</h1>
        <h2>Wildlife Video Classifier</h2>
        <p><b>Version:</b> 1.0.0 Beta</p>
        <hr>
        <p><b>Created by:</b> Nathan Bluto</p>
        <p><b>Data from:</b> The Gray Wolf Research Project</p>
        <p><b>Facilitated by:</b> Dr. Ausband</p>
        <hr>
        <p>An AI-powered tool for automated species identification in trail camera footage.</p>
        <p>Focused on Idaho fauna with expanding species coverage.</p>
        <hr>
        <p style="font-size: small; color: {THEME_CONFIG['text_secondary']};">
        Built with PyQt6 and powered by YOLO deep learning models.
        </p>
        </div>
        """
        
        QMessageBox.about(self, "About WolfVue", about_text)
    
    # ========================================================================
    # SETTINGS PERSISTENCE METHODS
    # ========================================================================
    
    def loadSettings(self):
        """Load settings from persistent storage."""
        settings = QSettings()
        
        # Load processing settings
        self.settings.input_folder = settings.value("processing/input_folder", "")
        self.settings.output_folder = settings.value("processing/output_folder", "")
        self.settings.model_path = settings.value("processing/model_path", "")
        self.settings.config_file = settings.value("processing/config_file", "")
        
        # Load advanced settings with defaults
        self.settings.confidence_threshold = float(settings.value(
            "processing/confidence_threshold", DEFAULT_PROCESSING_SETTINGS['confidence_threshold']
        ))
        
        # Load UI settings
        processing_mode = settings.value("ui/processing_mode", "Balanced")
        
        # Update UI with loaded settings
        self.input_folder_edit.setText(self.settings.input_folder)
        self.output_folder_edit.setText(self.settings.output_folder)
        self.model_path_edit.setText(self.settings.model_path)
        self.config_file_edit.setText(self.settings.config_file)
        self.confidence_slider.setValue(int(self.settings.confidence_threshold * 100))
        
        # Find and set processing mode
        index = self.processing_mode_combo.findText(processing_mode)
        if index >= 0:
            self.processing_mode_combo.setCurrentIndex(index)
        
        # Load window geometry
        geometry = settings.value("window/geometry")
        if geometry:
            self.restoreGeometry(geometry)
        
        # Load window state
        window_state = settings.value("window/state")
        if window_state:
            self.restoreState(window_state)
        
    def saveSettings(self):
        """Save settings to persistent storage."""
        settings = QSettings()
        
        # Save processing settings
        settings.setValue("processing/input_folder", self.settings.input_folder)
        settings.setValue("processing/output_folder", self.settings.output_folder)
        settings.setValue("processing/model_path", self.settings.model_path)
        settings.setValue("processing/config_file", self.settings.config_file)
        settings.setValue("processing/confidence_threshold", self.settings.confidence_threshold)
        
        # Save UI settings
        settings.setValue("ui/processing_mode", self.processing_mode_combo.currentText())
        
        # Save window geometry and state
        settings.setValue("window/geometry", self.saveGeometry())
        settings.setValue("window/state", self.saveState())
        
        # Save theme settings
        settings.setValue("theme/config", json.dumps(THEME_CONFIG))
        settings.setValue("font/config", json.dumps(FONT_CONFIG))
        
    def closeEvent(self, event):
        """Handle application close event."""
        if self.is_processing():
            reply = QMessageBox.question(
                self, "Processing in Progress",
                "Processing is currently running. Do you want to stop and exit?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                self.stop_processing()
                self.saveSettings()
                event.accept()
            else:
                event.ignore()
        else:
            self.saveSettings()
            event.accept()

# ============================================================================
# APPLICATION ENTRY POINT
# ============================================================================

def main():
    """
    Main application entry point.
    
    This function sets up the Qt application and creates the main window.
    To customize the application startup:
    1. Modify the application properties below
    2. Add command-line argument parsing if needed
    3. Add splash screen or initialization dialogs
    """
    import argparse
    
    # Parse command line arguments
    parser = argparse.ArgumentParser(description="WolfVue: Wildlife Video Classifier")
    parser.add_argument("--demo", action="store_true", help="Run in demo mode")
    parser.add_argument("--theme", choices=["light", "dark", "nature"], help="Set theme")
    parser.add_argument("--input", help="Input folder path")
    parser.add_argument("--output", help="Output folder path")
    args = parser.parse_args()
    
    # Create the Qt application
    app = QApplication(sys.argv)
    
    # Set application properties
    app.setApplicationName("WolfVue")
    app.setApplicationDisplayName("WolfVue: Wildlife Video Classifier")
    app.setApplicationVersion("1.0.0")
    app.setOrganizationName("Gray Wolf Research Project")
    app.setOrganizationDomain("wolfresearch.org")
    
    # Set application-wide styling
    app.setStyle("Fusion")  # Use Fusion style for consistency across platforms
    
    # Note: High DPI scaling is automatic in PyQt6, no manual setup needed
    
    # Create and configure the main window
    window = WolfVueMainWindow()
    
    # Apply command line arguments
    if args.theme:
        window.apply_preset_theme(args.theme)
    
    if args.input:
        window.input_folder_edit.setText(args.input)
        window.settings.input_folder = args.input
    
    if args.output:
        window.output_folder_edit.setText(args.output)
        window.settings.output_folder = args.output
    
    # Show the main window
    window.show()
    
    # Start the application event loop
    sys.exit(app.exec())

if __name__ == "__main__":
    main()